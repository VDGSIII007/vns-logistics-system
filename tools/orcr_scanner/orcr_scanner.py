"""
OR/CR OCR Scanner — VNS Trucking
Scans paper OR/CR documents, OCRs them, matches plate numbers from Central
Dispatch API, extracts registration fields, and outputs a review-ready Excel file.
"""
from __future__ import annotations

import os
import re
import sys
import logging
from pathlib import Path
from datetime import datetime

import requests
import pandas as pd
import fitz  # PyMuPDF
import pytesseract
import cv2
import numpy as np
from PIL import Image
from rapidfuzz import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ── Configuration ─────────────────────────────────────────────────────────────
ORCR_FOLDER = r"C:\Users\User\Documents\AllTrucksORCR\ORCR NEW UPDATED"
OUTPUT_FILE = r"output\truck_master_extracted.xlsx"
CENTRAL_API_URL = (
    "https://script.google.com/macros/s/"
    "AKfycbwkA_gMbqPvtW3kEDsCKAkgylrakQwRHlPNPYENT2GYvjH1AGAsmusUuPUvWrB_KakH"
    "/exec?action=getDispatchDashboard"
)
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

SUPPORTED_EXT   = {".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".tif", ".bmp"}
FUZZY_THRESHOLD = 86
OCR_DPI_SCALE   = 4
LOW_CONFIDENCE  = 60
OCR_CONFIG      = "--oem 3 --psm 6"
ROTATIONS       = (0, 90, 180, 270)
NON_ORCR_FILENAME_KEYWORDS = (
    "APPLICATION",
    "AUTHORITY",
    "AUTHO",
    "DECISION",
    "DROPPING",
    "FRANCHISE",
    "INSURANCE",
    "NOTICE",
)

# ── Logging ───────────────────────────────────────────────────────────────────
os.makedirs("output", exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("output/orcr_scan.log", encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

# ── Field extraction patterns ─────────────────────────────────────────────────
# Keys must match MAIN_COLS exactly.
FIELD_PATTERNS: dict[str, list[str]] = {
    "Registered Owner": [
        r"REGISTERED\s+OWNER[:\s]+(.+?)(?=\n|MAKE|ADDRESS|$)",
        r"OWNER[:\s]+(.+?)(?=\n|MAKE|ADDRESS|$)",
    ],
    "Make / Brand": [
        r"(?:MAKE\s*/\s*BRAND|MAKE|BRAND)[:\s]+(.+?)(?=\n|SERIES|MODEL|YEAR|$)",
    ],
    "Series / Model": [
        r"(?:SERIES\s*/\s*MODEL|SERIES|MODEL)[:\s]+(.+?)(?=\n|YEAR|BODY|$)",
    ],
    "Year Model": [
        r"(?:YEAR\s+MODEL|YEAR\s+OF\s+MANUF\.?|YEAR)[:\s]+(\d{4})",
    ],
    "Body Type": [
        r"BODY\s+TYPE[:\s]+(.+?)(?=\n|FUEL|COLOU?R|ENGINE|$)",
    ],
    "Fuel Type": [
        r"FUEL\s+TYPE[:\s]+(.+?)(?=\n|COLOU?R|ENGINE|$)",
        r"FUEL[:\s]+(.+?)(?=\n|COLOU?R|ENGINE|$)",
    ],
    "Color": [
        r"(?:COLOU?R)[:\s]+(.+?)(?=\n|ENGINE|CHASSIS|MV FILE|$)",
    ],
    "Engine No.": [
        r"ENGINE\s+(?:NO\.?|NUMBER)[:\s]+([A-Z0-9\-]+)",
    ],
    "Chassis No.": [
        r"CHASSIS\s+(?:NO\.?|NUMBER)[:\s]+([A-Z0-9\-]+)",
    ],
    "MV File No.": [
        r"MV\s+FILE\s+(?:NO\.?|NUMBER)[:\s]+([A-Z0-9\-]+)",
        r"MVFILE[:\s]+([A-Z0-9\-]+)",
    ],
    "CR No.": [
        r"CR\s+(?:NO\.?|NUMBER)[:\s]+([A-Z0-9\-]+)",
        r"CERTIF(?:ICATE)?\s+OF\s+REG[A-Z.]*\s+(?:NO\.?|#)\s*([A-Z0-9\-]+)",
    ],
    "OR No.": [
        r"OR\s+(?:NO\.?|NUMBER)[:\s]+([A-Z0-9\-]+)",
        r"OFFICIAL\s+RECEIPT\s+(?:NO\.?|#)\s*([A-Z0-9\-]+)",
    ],
    "Gross Weight": [
        r"GROSS\s+(?:VEHICLE\s+)?WEIGHT[:\s]+([\d,\.]+(?:\s*KGS?|LBS?)?)",
        r"GVW[:\s]+([\d,\.]+(?:\s*KGS?|LBS?)?)",
    ],
    "Net Capacity": [
        r"NET\s+CAPACITY[:\s]+([\d,\.]+(?:\s*KGS?|LBS?)?)",
        r"NET\s+CAP(?:ACITY)?[:\s]+([\d,\.]+(?:\s*KGS?|LBS?)?)",
    ],
    "Registration Date": [
        r"(?:REGISTRATION\s+DATE|REG(?:ISTRATION)?\.\s*DATE|DATE\s+OF\s+REG(?:ISTRATION)?)"
        r"[:\s]+([A-Z0-9\s\-/,\.]+?)(?=\n|EXPIR|VALID|$)",
    ],
    "Registration Expiry": [
        r"(?:EXPIR(?:Y|ATION)\s+DATE|VALID\s+UNTIL|EXPIRATION|EXPIRES?)"
        r"[:\s]+([A-Z0-9\s\-/,\.]+?)(?=\n|$)",
    ],
}

# Fields that are human-readable text (title-cased); others kept as-is
TEXT_FIELDS = {"Registered Owner", "Make / Brand", "Series / Model", "Body Type", "Fuel Type", "Color"}

# ── Excel column definitions ──────────────────────────────────────────────────
MAIN_COLS = [
    "Plate Number", "Matched File Name", "Matched File Path", "Match Method", "Match Score",
    "Registered Owner", "Make / Brand", "Series / Model", "Year Model", "Body Type",
    "Fuel Type", "Color", "Engine No.", "Chassis No.", "MV File No.", "CR No.", "OR No.",
    "Gross Weight", "Net Capacity", "Registration Date", "Registration Expiry",
    "Raw Text Preview", "Confidence", "Needs Review", "Notes",
    "groupCategory", "driverName", "helperName", "imei", "status",
]
MISSING_COLS   = ["Plate Number", "groupCategory", "driverName", "helperName", "imei", "status", "Notes"]
UNMATCHED_COLS = ["File Name", "File Path", "Confidence", "Raw Text Preview", "Notes"]
DEBUG_COLS     = ["File Name", "File Path", "Matched Plate", "Match Method", "Match Score", "Confidence", "Raw Text Preview"]


# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize_plate(plate: str) -> str:
    """Remove spaces and dashes, uppercase — used for comparison only."""
    return re.sub(r"[\s\-]", "", str(plate).upper())


def clean_ocr_text(text: str) -> str:
    """Normalize common OCR artifacts without destroying LTO form line layout."""
    text = text.replace("|", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def clean_field_value(value: str, title_case: bool = False) -> str:
    value = re.sub(r"\s+", " ", str(value).strip(" :-_.,'\""))
    value = re.sub(r"\bN/?A\b", "N/A", value, flags=re.IGNORECASE)
    if title_case:
        value = re.sub(r"^[^A-Za-z]+", "", value)
        value = re.sub(r"^[A-Z]\s+(?=[A-Z]{2,})", "", value, flags=re.IGNORECASE)
        value = re.sub(r"\s+(?:EE|ES|BE|AC|X)$", "", value, flags=re.IGNORECASE)
        value = re.sub(r"\s+[A-Z]$", "", value, flags=re.IGNORECASE)
    return value.title() if title_case and value and value != "N/A" else value.upper()


def filename_plate_match(file_path: str, plate_list: list[str]) -> tuple[str | None, str | None, int]:
    """
    Fast pre-OCR match. This intentionally only looks at the filename so we can
    skip authority/decision/application files that do not belong to roster plates.
    """
    if not plate_list:
        return None, None, 0

    stem_clean = normalize_plate(Path(file_path).stem)
    norm_plates = [normalize_plate(p) for p in plate_list]

    for plate, norm in zip(plate_list, norm_plates):
        if norm and (norm in stem_clean or stem_clean == norm):
            return plate, "filename_exact", 100

    result = process.extractOne(stem_clean, norm_plates, scorer=fuzz.partial_ratio)
    if result and result[1] >= FUZZY_THRESHOLD:
        idx = norm_plates.index(result[0])
        return plate_list[idx], "fuzzy_filename", int(round(result[1]))

    return None, None, 0


# ── Central Dispatch API ──────────────────────────────────────────────────────
def fetch_plates_from_api() -> list[dict]:
    """
    Pull truck roster from Central Dispatch API.
    Returns a list of dicts with plateNumber, groupCategory, driverName,
    helperName, imei, status.  Returns [] on any failure.
    """
    try:
        log.info("Connecting to Central Dispatch API …")
        resp = requests.get(CENTRAL_API_URL, timeout=30)
        resp.raise_for_status()
        data = resp.json()
    except requests.RequestException as exc:
        log.error(f"API request failed: {exc}")
        return []
    except ValueError as exc:
        log.error(f"API response is not valid JSON: {exc}")
        return []

    # Normalise varying response shapes
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        items = (
            data.get("trucks")
            or data.get("data")
            or data.get("records")
            or data.get("rows")
            or []
        )
        if not items and data:
            first = list(data.values())
            items = first[0] if first and isinstance(first[0], list) else first
    else:
        items = []

    trucks = []
    for item in items:
        if not isinstance(item, dict):
            continue
        plate = (
            item.get("plateNumber") or item.get("plate_number")
            or item.get("plate") or item.get("Plate Number")
            or item.get("PLATE") or ""
        ).strip().upper()
        if plate:
            trucks.append({
                "plateNumber":   plate,
                "groupCategory": item.get("groupCategory", item.get("group",    "")),
                "driverName":    item.get("driverName",    item.get("driver",   "")),
                "helperName":    item.get("helperName",    item.get("helper",   "")),
                "imei":          item.get("imei",          item.get("IMEI",     "")),
                "status":        item.get("status",        item.get("Status",   "")),
            })

    log.info(f"  → {len(trucks)} plate numbers fetched.")
    return trucks


# ── Image preprocessing & OCR ─────────────────────────────────────────────────
def preprocess(img: Image.Image) -> Image.Image:
    """Denoise + Otsu-threshold a PIL image to improve Tesseract accuracy."""
    arr = np.array(img.convert("RGB"))
    gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
    denoised = cv2.fastNlMeansDenoising(gray, h=10)
    _, thresh = cv2.threshold(denoised, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    return Image.fromarray(thresh)


def ocr_page(img: Image.Image) -> tuple[str, float]:
    """Run Tesseract and preserve line breaks from the LTO form layout."""
    try:
        data = pytesseract.image_to_data(
            img,
            output_type=pytesseract.Output.DICT,
            lang="eng",
            config=OCR_CONFIG,
        )
    except Exception as exc:
        log.warning(f"  Tesseract error: {exc}")
        return "", 0.0

    lines: dict[tuple[int, int, int], list[tuple[int, str]]] = {}
    confs: list[float] = []

    for i, word in enumerate(data["text"]):
        word = str(word).strip()
        if not word:
            continue
        try:
            conf = float(data["conf"][i])
        except (TypeError, ValueError):
            conf = -1
        if conf > 0:
            confs.append(conf)

        key = (data["block_num"][i], data["par_num"][i], data["line_num"][i])
        lines.setdefault(key, []).append((data["left"][i], word))

    text_lines = []
    for key in sorted(lines):
        words = [word for _, word in sorted(lines[key], key=lambda item: item[0])]
        text_lines.append(" ".join(words))

    text = clean_ocr_text("\n".join(text_lines))
    avg_conf = round(sum(confs) / len(confs), 1) if confs else 0.0
    return text, avg_conf


def ocr_page_best_rotation(img: Image.Image) -> tuple[str, float, int]:
    """Try all likely scan rotations and keep the most useful LTO OCR result."""
    best_text = ""
    best_conf = 0.0
    best_rotation = 0
    best_score = -1.0

    for rotation in ROTATIONS:
        candidate = img.rotate(rotation, expand=True) if rotation else img
        text, conf = ocr_page(candidate)
        upper = text.upper()
        keyword_hits = sum(
            phrase in upper
            for phrase in (
                "CERTIFICATE OF REGISTRATION",
                "OFFICIAL RECEIPT",
                "LAND TRANSPORTATION OFFICE",
                "PLATE NO",
                "ENGINE NO",
                "CHASSIS NO",
                "OR NO",
                "PAYMENT DETAILS",
            )
        )
        score = conf + (keyword_hits * 8) + min(len(text) / 500, 8)
        if score > best_score:
            best_text = text
            best_conf = conf
            best_rotation = rotation
            best_score = score

    return best_text, best_conf, best_rotation


def extract_text(file_path: str) -> tuple[str, float]:
    """
    OCR a PDF or image file.  PDFs are rendered page-by-page at OCR_DPI_SCALE.
    Returns (full_text, average_confidence).
    """
    ext = Path(file_path).suffix.lower()
    texts: list[str] = []
    confs: list[float] = []

    try:
        if ext == ".pdf":
            doc = fitz.open(file_path)
            mat = fitz.Matrix(OCR_DPI_SCALE, OCR_DPI_SCALE)
            for page_no, page in enumerate(doc, 1):
                pix = page.get_pixmap(matrix=mat)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                img = preprocess(img)
                t, c, rotation = ocr_page_best_rotation(img)
                texts.append(f"--- PAGE {page_no} ROTATION {rotation} ---\n{t}")
                confs.append(c)
            doc.close()
        else:
            img = Image.open(file_path)
            img = preprocess(img)
            t, c, rotation = ocr_page_best_rotation(img)
            texts.append(f"--- PAGE 1 ROTATION {rotation} ---\n{t}")
            confs.append(c)
    except Exception as exc:
        log.error(f"  Cannot process {Path(file_path).name}: {exc}")

    full_text = "\n".join(texts)
    avg_conf  = round(sum(confs) / len(confs), 1) if confs else 0.0
    return full_text, avg_conf


# ── Plate matching ────────────────────────────────────────────────────────────
def match_to_plate(
    file_path: str,
    ocr_text: str,
    plate_list: list[str],
) -> tuple[str | None, str | None, int]:
    """
    Match a file to a plate number using four strategies in priority order:
      1. filename_exact  — plate token found verbatim in the filename
      2. ocr_exact       — plate token found verbatim in OCR text
      3. fuzzy_filename  — rapidfuzz partial_ratio on the filename
      4. fuzzy_ocr       — plate-like tokens in OCR text matched fuzzily
    Returns (matched_plate, method, score) or (None, None, 0).
    """
    if not plate_list:
        return None, None, 0

    plate, method, score = filename_plate_match(file_path, plate_list)
    if plate:
        return plate, method, score

    norm_plates = [normalize_plate(p) for p in plate_list]

    if ocr_text:
        text_nospace = re.sub(r"[\s\-]", "", ocr_text.upper())
        for plate, norm in zip(plate_list, norm_plates):
            if norm in text_nospace:
                return plate, "ocr_exact", 95

    if ocr_text:
        candidates = re.findall(r"[A-Z]{2,3}[\s\-]?\d{3,4}[A-Z]?", ocr_text.upper())
        for cand in candidates:
            cand_clean = normalize_plate(cand)
            res = process.extractOne(cand_clean, norm_plates, scorer=fuzz.ratio)
            if res and res[1] >= 85:
                idx = norm_plates.index(res[0])
                return plate_list[idx], "fuzzy_ocr", res[1]

    return None, None, 0


# ── Field extraction ──────────────────────────────────────────────────────────
def extract_fields(text: str) -> dict[str, str]:
    """Extract OR/CR registration fields from line-preserved OCR text."""
    upper = clean_ocr_text(text.upper())
    result: dict[str, str] = {field: "" for field in FIELD_PATTERNS}
    bad_values = {
        "",
        "=",
        "-",
        "CHASSIS",
        "CHASSIS NO",
        "VIN",
        "SERIES",
        "SERIES GROSS WEIGHT NET WEIGHT",
        "BODY NO",
    }

    def assign(field: str, value: str, title: bool | None = None) -> None:
        value = clean_field_value(value, field in TEXT_FIELDS if title is None else title)
        current = result.get(field, "")
        if value and (not current or current in bad_values or len(current) <= 2):
            result[field] = value

    def grab(field: str, *patterns: str, title: bool | None = None) -> None:
        if result.get(field):
            return
        for pattern in patterns:
            match = re.search(pattern, upper, re.MULTILINE | re.DOTALL)
            if match:
                assign(field, match.group(1), title)
                return

    # LTO Certificate of Registration page.
    grab(
        "Registered Owner",
        r"OWNER'?S\s+NAME\s*\n+(.+?)(?=\nOWNER'?S\s+ADDRESS|\nENCUMBERED|\nDETAILS|\Z)",
        r"OWNER'?S\s+NAME.*?\n+(.+?)(?=\n.*OWNER'?S\s+ADDRESS|\n.*ENCUMBERED|\n.*DETAILS|\Z)",
        r"REGISTERED\s+OWNER\s*\n+(.+?)(?=\nADDRESS|\nMAKE|\Z)",
        title=True,
    )
    grab("Make / Brand", r"^MAKE\s*/?\s*BRAND\s*\n+([A-Z0-9 /.-]+?)(?=\n|$)", title=True)
    grab("Series / Model", r"^SERIES\s*\n+([A-Z0-9 /.-]+?)(?=\n|$)", title=True)
    grab("Year Model", r"YEAR\s+MODEL(?:\s*\([^)]*\))?\s*\n+(\d{4})")
    grab("Body Type", r"^BODY\s+TYPE\s*\n+([A-Z0-9 /.-]+?)(?=\n|$)", title=True)
    grab("Fuel Type", r"^TYPE\s+OF\s+FUEL\s*\n+([A-Z0-9 /.-]+?)(?=\n|$)", title=True)
    grab("Color", r"^COLOR\s*\n+([A-Z0-9 /.-]+?)(?=\n|$)", title=True)
    grab("Engine No.", r"ENGINE\s+NO\.?\s*\n+([A-Z0-9\-]+)")
    grab("Chassis No.", r"CHASSIS\s+NO\.?\s*\n+([A-Z0-9\-]+)")
    grab("MV File No.", r"\bFILE\s+NO\.?\s*\n+([A-Z0-9\-]+)")
    grab("CR No.", r"\bCR\s+NO\.?\s*[-:.]?\s*([A-Z0-9\-]+)")
    grab("OR No.", r"\b[O0]\.?\s*R\.?\s+NO\.?\s*\n+([A-Z0-9\-]+)", r"\bOR\s+NO\.?\s*[:.]?\s*([A-Z0-9\-]+)")
    grab("Gross Weight", r"GROSS\s+WEIGHT\s*\n+([\d,\.]+)")
    grab("Net Capacity", r"NET\s+(?:WEIGHT|CAPACITY)\s*\n+([\d,\.]+)")
    grab("Registration Date", r"\bO\.?\s*R\.?\s+DATE\s*\n+([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})")
    grab(
        "Registration Expiry",
        r"VALID\s+UNTIL[\s\S]{0,60}?([0-9]{1,2}/[0-9]{2,4})",
        r"VALID\s+UNTIL\s+([0-9]{1,2}/[0-9]{2,4})",
        r"VALID\s+UNTIL\s+([A-Z0-9 /\-]+?)(?=\s+AND\s+DUE|\n|$)",
    )

    # Table-row fallback for OCR that reads labels in one line and values in the next.
    lines = [line.strip() for line in upper.splitlines() if line.strip()]
    for idx, line in enumerate(lines[:-1]):
        nearby = lines[idx + 1:idx + 5]
        next_line = lines[idx + 1]
        tokens = next_line.split()
        if "PLATE NO" in line and "ENGINE NO" in line and "CHASSIS NO" in line:
            value_line = next((near for near in nearby if re.search(r"\b[A-Z]{2,3}\d{3,4}\b", near)), next_line)
            value_tokens = value_line.split()
            plate_pos = next((pos for pos, tok in enumerate(value_tokens) if re.fullmatch(r"[A-Z]{2,3}\d{3,4}", tok)), -1)
            if plate_pos >= 0 and len(value_tokens) > plate_pos + 2:
                assign("Engine No.", value_tokens[plate_pos + 1])
                chassis = next((tok for tok in value_tokens[plate_pos + 2:] if re.fullmatch(r"[A-Z0-9][A-Z0-9\-]{5,}", tok)), "")
                assign("Chassis No.", chassis)
        elif "FILE NO" in line and "MAKE/BRAND" in line:
            value_line = next((near for near in nearby if re.search(r"\b\d{10,}\b", near)), next_line)
            value_tokens = value_line.split()
            file_no = next((tok for tok in value_tokens if re.fullmatch(r"\d{10,}", tok)), "")
            if file_no:
                assign("MV File No.", file_no)
            brand = next(
                (
                    tok for tok in reversed(value_tokens)
                    if len(tok) > 2 and tok not in {"TRUCK", "BUS", "MOTOR", "VEHICLE"}
                    and not re.fullmatch(r"\d+", tok)
                ),
                "",
            )
            assign("Make / Brand", brand, title=True)
        elif "COLOR" in line and "TYPE OF FUEL" in line:
            value_line = next((near for near in nearby if any(fuel in near.split() for fuel in ("DIESEL", "GAS", "GASOLINE", "ELECTRIC", "HYBRID"))), next_line)
            value_tokens = value_line.split()
            color_idx = 1 if value_tokens and re.fullmatch(r"\d+", value_tokens[0]) and len(value_tokens) > 1 else 0
            if value_tokens:
                assign("Color", value_tokens[color_idx], title=True)
            fuel = ""
            for fuel_type in ("DIESEL", "GAS", "GASOLINE", "ELECTRIC", "HYBRID"):
                if fuel_type in value_tokens:
                    fuel = fuel_type
                    break
            assign("Fuel Type", fuel or (value_tokens[-1] if value_tokens else ""), title=True)
        elif "BODY TYPE" in line and "GROSS WEIGHT" in line:
            value_lines = [near for near in nearby if re.search(r"\b\d{4,6}\b", near)]
            value_line = " ".join(value_lines[:2]) if value_lines else next_line
            value_tokens = value_line.split()
            numeric_positions = [pos for pos, tok in enumerate(value_tokens) if re.fullmatch(r"\d{4,6}", tok)]
            if numeric_positions:
                assign("Gross Weight", value_tokens[numeric_positions[0]])
            if len(numeric_positions) > 1:
                assign("Net Capacity", value_tokens[numeric_positions[1]])
            if numeric_positions:
                series_pos = numeric_positions[0] - 1
                assign("Series / Model", value_tokens[series_pos], title=True)
                body = " ".join(value_tokens[:series_pos])
                assign("Body Type", body, title=True)
                body_line = next((near for near in nearby if not re.search(r"\d{4,6}", near) and re.search(r"\b(TRUCK|BUS|VAN)\b", near)), "")
                if body_line:
                    assign("Body Type", body_line, title=True)
        elif "YEAR MODEL" in line and len(tokens) >= 1 and not result["Year Model"]:
            value_line = next((near for near in nearby if re.search(r"\b(19|20)\d{2}\b", near)), next_line)
            year = next((tok for tok in value_line.split() if re.fullmatch(r"(19|20)\d{2}", tok)), "")
            assign("Year Model", year)
        elif "O.R. NO" in line or "0.R. NO" in line or "OR NO" in line:
            value_line = next((near for near in nearby if re.search(r"\d{1,2}/\d{1,2}/\d{2,4}", near)), next_line)
            value_tokens = value_line.split()
            or_no = next((tok for tok in value_tokens if re.fullmatch(r"[A-Z0-9\-]{8,}", tok)), "")
            date = next((tok for tok in value_tokens if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{2,4}", tok)), "")
            assign("OR No.", or_no)
            assign("Registration Date", date)

    # Older generic patterns as a last resort.
    for field, patterns in FIELD_PATTERNS.items():
        current = result.get(field, "")
        if current and current not in bad_values and not current.startswith("SERIES GROSS"):
            continue
        for pattern in patterns:
            m = re.search(pattern, upper, re.MULTILINE)
            if m:
                val = m.group(1).strip()
                result[field] = clean_field_value(val, field in TEXT_FIELDS)
                break
    return result


# ── File scanner ──────────────────────────────────────────────────────────────
def scan_folder() -> list[Path]:
    root = Path(ORCR_FOLDER)
    if not root.exists():
        log.error(f"ORCR folder not found: {ORCR_FOLDER}")
        return []
    files = [p for p in root.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_EXT]
    log.info(f"Found {len(files)} file(s) in {root}")
    return files


def select_plate_named_files(files: list[Path], plate_list: list[str]) -> list[tuple[Path, str, str, int]]:
    """Keep only files whose names match roster plates, choosing the best file per plate."""
    best_by_plate: dict[str, tuple[tuple[int, int, int, str], Path, str, int]] = {}

    for fp in files:
        plate, method, score = filename_plate_match(str(fp), plate_list)
        if not plate:
            continue

        stem = fp.stem.upper()
        stem_clean = normalize_plate(stem)
        plate_clean = normalize_plate(plate)
        keyword_penalty = 1 if any(keyword in stem for keyword in NON_ORCR_FILENAME_KEYWORDS) else 0
        exact_name_penalty = 0 if stem_clean == plate_clean else 1
        duplicate_penalty = 1 if re.search(r"\(\d+\)$", stem) else 0
        rank = (keyword_penalty, exact_name_penalty, duplicate_penalty, stem)

        current = best_by_plate.get(plate)
        if current is None or rank < current[0]:
            best_by_plate[plate] = (rank, fp, method or "", score)

    selected = [
        (fp, plate, method, score)
        for plate, (_, fp, method, score) in best_by_plate.items()
    ]
    selected.sort(key=lambda item: item[1])
    log.info(f"Selected {len(selected)} plate-matched file(s); skipped non-matching and duplicate files.")
    return selected


# ── Excel output ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
REVIEW_FILL  = PatternFill("solid", fgColor="FFEB9C")   # amber — needs review
MISSING_FILL = PatternFill("solid", fgColor="FCE4D6")   # salmon — missing ORCR


def _to_df(rows: list[dict], cols: list[str]) -> pd.DataFrame:
    return pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)


def _style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    for ws in wb.worksheets:
        # Header row
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"

        # Auto column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        # Row highlights
        if ws.title == "Truck_Master_Extracted":
            nr_col = next((c.column for c in ws[1] if c.value == "Needs Review"), None)
            if nr_col:
                for row in ws.iter_rows(min_row=2):
                    if row[nr_col - 1].value == "Yes":
                        for cell in row:
                            cell.fill = REVIEW_FILL

        elif ws.title == "Missing_ORCR":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = MISSING_FILL

    wb.save(path)


def write_excel(
    extracted:  list[dict],
    missing:    list[dict],
    unmatched:  list[dict],
    debug:      list[dict],
) -> None:
    out = Path(OUTPUT_FILE)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        _to_df(extracted,  MAIN_COLS).to_excel(writer, sheet_name="Truck_Master_Extracted", index=False)
        _to_df(missing,    MISSING_COLS).to_excel(writer, sheet_name="Missing_ORCR",         index=False)
        _to_df(unmatched,  UNMATCHED_COLS).to_excel(writer, sheet_name="Unmatched_Files",    index=False)
        _to_df(debug,      DEBUG_COLS).to_excel(writer, sheet_name="Raw_Debug",              index=False)

    _style_workbook(out)
    log.info(f"Excel saved → {out.resolve()}")


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    log.info("=" * 60)
    log.info("OR/CR OCR Scanner — VNS Trucking")
    log.info(f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    log.info("=" * 60)

    # 1. Fetch plate roster from Central Dispatch API
    truck_data = fetch_plates_from_api()
    plate_list = [t["plateNumber"] for t in truck_data]
    plate_meta = {t["plateNumber"]: t for t in truck_data}
    if not plate_list:
        log.warning("No plate numbers fetched — files will be OCR'd but not matched.")

    # 2. Discover all OR/CR files
    files = scan_folder()
    if not files:
        log.error("No files found. Exiting.")
        sys.exit(1)
    plate_files = select_plate_named_files(files, plate_list)

    extracted: list[dict] = []
    unmatched: list[dict] = []
    debug:     list[dict] = []

    # 3. OCR + extract each selected plate file
    for i, (fp, plate, method, score) in enumerate(plate_files, 1):
        name = fp.name
        log.info(f"[{i:>3}/{len(plate_files)}] {name} -> {plate}")

        raw_text, confidence = extract_text(str(fp))
        ocr_plate, ocr_method, ocr_score = match_to_plate(str(fp), raw_text, plate_list)
        if ocr_plate == plate and ocr_method and ocr_score > score:
            method, score = ocr_method, ocr_score

        debug.append({
            "File Name":        name,
            "File Path":        str(fp),
            "Matched Plate":    plate or "",
            "Match Method":     method or "",
            "Match Score":      score,
            "Confidence":       confidence,
                "Raw Text Preview": raw_text[:500],
        })

        fields = extract_fields(raw_text)
        meta   = plate_meta.get(plate, {})

        notes: list[str] = []
        if confidence < LOW_CONFIDENCE:  notes.append("Low OCR confidence")
        if score < 85:                   notes.append(f"Low match score ({score})")
        if not any(fields.values()):     notes.append("No fields extracted")
        needs_review = "Yes" if notes else "No"

        extracted.append({
            "Plate Number":      plate,
            "Matched File Name": name,
            "Matched File Path": str(fp),
            "Match Method":      method,
            "Match Score":       score,
            **fields,
            "Raw Text Preview":  raw_text[:300],
            "Confidence":        confidence,
            "Needs Review":      needs_review,
            "Notes":             "; ".join(notes),
            "groupCategory":     meta.get("groupCategory", ""),
            "driverName":        meta.get("driverName",    ""),
            "helperName":        meta.get("helperName",    ""),
            "imei":              meta.get("imei",          ""),
            "status":            meta.get("status",        ""),
        })

    # 4. Identify plates with no matching file
    matched_plates = {r["Plate Number"] for r in extracted}
    missing: list[dict] = [
        {
            "Plate Number":  p,
            "groupCategory": plate_meta[p].get("groupCategory", ""),
            "driverName":    plate_meta[p].get("driverName",    ""),
            "helperName":    plate_meta[p].get("helperName",    ""),
            "imei":          plate_meta[p].get("imei",          ""),
            "status":        plate_meta[p].get("status",        ""),
            "Notes":         "No matching OR/CR file found",
        }
        for p in plate_list if p not in matched_plates
    ]

    # 5. Write Excel
    write_excel(extracted, missing, unmatched, debug)

    log.info("─" * 60)
    log.info(f"  Matched:   {len(extracted)}")
    log.info(f"  Missing:   {len(missing)}")
    log.info(f"  Unmatched: {len(unmatched)}")
    log.info(f"  Output:    {Path(OUTPUT_FILE).resolve()}")
    log.info("─" * 60)


if __name__ == "__main__":
    main()
